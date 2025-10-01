import pandas as pd
import uuid
from datetime import date
from utils import to_decimal, find_zodiac, find_nakshatra, generate_full_dasha
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from NAKSHATRA_TABLE import NAKSHATRAS

PLANET_STYLE = {
    "Ketu":    {"symbol": "☋", "color": "494529"},
    "Venus":   {"symbol": "♀", "color": "BFBFBF"},
    "Sun":     {"symbol": "☉", "color": "FF99CC"},
    "Moon":    {"symbol": "☽", "color": "FFCCFF"},
    "Mars":    {"symbol": "♂", "color": "FF0000"},
    "Rahu":    {"symbol": "☊", "color": "F2F2F2"},
    "Jupiter": {"symbol": "♃", "color": "00B050"},
    "Saturn":  {"symbol": "♄", "color": "FFFF00"},
    "Mercury": {"symbol": "☿", "color": "00B0F0"},
}

def main():
    print("=== Vimshottari Dasha Calculator ===")

    # --- 사용자 입력 ---
    zodiac = input("Zodiac sign (예: Taurus): ").capitalize()
    degree = int(input("Degree (0-29): "))
    minute = int(input("Minute (0-59): "))

    birth_year = int(input("Birth year (YYYY): "))
    birth_month = int(input("Birth month (MM): "))
    birth_day = int(input("Birth day (DD): "))

    birth_date = date(birth_year, birth_month, birth_day)

    # --- decimal 변환 ---
    decimal = to_decimal(zodiac, degree, minute)

    # --- zodiac & nakshatra 찾기 ---
    zod = find_zodiac(decimal)
    nak, ruler = find_nakshatra(decimal)

    print(f"\n[입력값 정리]")
    print(f"  Zodiac : {zod}")
    print(f"  Nakshatra : {nak} (Ruler: {ruler})")


    # 낙샤트라 범위
    for info in NAKSHATRAS:
        if info["name"] == nak:
            total_minutes = (info["end"] - info["start"]) * 60
            current_minutes = (decimal - info["start"]) * 60
            # ✅ 이미 지난 분 = current_minutes (0~799)
            offset_minute = int(current_minutes)
            break

    # 전체 대샤 생성
    rows = generate_full_dasha(birth_date, ruler, offset_minute)


    # --- 날짜 문자열화 ---
    formatted_rows = []
    for row in rows:
        start_date, age, lv1, lv2, lv3 = row
        formatted_rows.append([
            start_date.strftime("%Y-%m-%d"),
            age,
            lv1,
            lv2,
            lv3
        ])

    df = pd.DataFrame(formatted_rows, columns=["Date", "Age", "Lv1", "Lv2", "Lv3"])

    # --- 랜덤 파일명 ---
    filename = f"dasha_output_{uuid.uuid4().hex[:6]}.xlsx"
    df.to_excel(filename, index=False)

    # --- 색상 스타일 적용 ---
    wb = load_workbook(filename)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=5):
        for cell in row:
            planet_name = str(cell.value).replace("☋ ", "").replace("♀ ", "").replace("☉ ", "") \
                                         .replace("☽ ", "").replace("♂ ", "").replace("☊ ", "") \
                                         .replace("♃ ", "").replace("♄ ", "").replace("☿ ", "")
            style = PLANET_STYLE.get(planet_name, None)
            if style:
                cell.value = f"{style['symbol']} {planet_name}"
                cell.fill = PatternFill(start_color=style["color"],
                                        end_color=style["color"],
                                        fill_type="solid")

    wb.save(filename)
    print(f"\n[엑셀 저장 완료] → {filename}")

if __name__ == "__main__":
    main()
